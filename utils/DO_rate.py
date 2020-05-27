#  分析DO完成率工具

import openpyxl
import pymysql
from utils.cosmos_db import cosmos

xlsx_path = '/Users/cai/cai/OtisCode/Doctor_Otis_API/'
date = '2020-5-24'
xlsx_name = date + '.xlsx'
mysql_host = 'do-api-database-vegctujtkpkfg.mysql.database.azure.com'
user_name = 'cai@do-api-database-vegctujtkpkfg'
password = '1Q2w3e4r'
xlsx = xlsx_path + xlsx_name

mysql_conn = pymysql.connect(host=mysql_host, user=user_name, password=password, database='do_api', charset="utf8")

cursor = mysql_conn.cursor()

wb = openpyxl.load_workbook(xlsx)
ws = wb[date]
ws.insert_cols(4)
ws.insert_cols(5)
ws.insert_cols(6)
ws.cell(row=1, column=4).value = '是否为IoT电梯'
ws.cell(row=1, column=5).value = '是否完成IoT DO'
ws.cell(row=1, column=6).value = 'IoT电梯是否离线3天以上'

rows = ws.max_row
for i in range(1, rows + 1):
    unit_number = ws.cell(row=i, column=3).value
    result = cosmos.query('COLLECTION_DSLOG_MASTER', fields=('UnitNumber',), query_params={'UnitNumber': unit_number})
    if result:
        print(unit_number)
        ws.cell(row=i, column=4).value = 'IoT'
        sql = """SELECT count(unit_number) from api_apirecord where unit_number='%s' 
                 and datetime between '%s 00:00:00' and '%s 23:59:59'""" % (unit_number, date, date)
        cursor.execute(sql)
        sql_results = cursor.fetchone()
        if sql_results[0] >= 1:
            print(unit_number, 'Y')
            ws.cell(row=i, column=5).value = 'Y'
        else:
            offline = cosmos.check_unit_offline_three_days(unit_number)
            if offline:
                ws.cell(row=i, column=6).value = 'Y'
                print(unit_number, '离线')


wb.save(xlsx_path + 'test1.xlsx')
cursor.close()
mysql_conn.close()
